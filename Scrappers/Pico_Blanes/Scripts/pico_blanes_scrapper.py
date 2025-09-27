import os
import json
import re
import time
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from bs4 import BeautifulSoup as bs
from typing import Optional, Tuple, Dict, List
from datetime import datetime
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Optional
from typing import Tuple, Optional
import glob
from pathlib import Path
import shutil
import csv
import requests
import sys
import os
import csv
import time
import shutil
import pandas as pd
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    WebDriverException,
    JavascriptException,
    TimeoutException,
)
import time
from urllib.parse import urljoin
from pathlib import Path

# Selenium / undetected-chromedriver
import undetected_chromedriver as uc
from selenium.common.exceptions import WebDriverException
from urllib.parse import urlparse, urlunparse

DEBUG = True
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = (
    SCRIPT_DIR.parent.parent
)  # .../Scrappers/Idealista/Scripts -> sube dos niveles
BASE_DIR = str((PROJECT_ROOT / "Pico_Blanes" / "Data").resolve())


def clear_browser_state(browser, *, clear_cache=True, clear_cookies=True):
    """
    Limpia caché y cookies vía CDP para cortar crecimiento de memoria.
    """
    try:
        if clear_cache:
            browser.execute_cdp_cmd("Network.clearBrowserCache", {})
        if clear_cookies:
            browser.delete_all_cookies()
    except Exception:
        pass


# -------- Helpers de rutas por 'busqueda' --------
def paths_for():
    ids_today = os.path.join(BASE_DIR, f"ids_today.csv")
    ids_yesterday = os.path.join(BASE_DIR, f"ids_yesterday.csv")
    ids_new = os.path.join(BASE_DIR, f"ids_new.csv")
    data_today = os.path.join(BASE_DIR, f"inmuebles_today.csv")
    data_new = os.path.join(BASE_DIR, f"inmuebles_new.csv")
    os.makedirs(BASE_DIR, exist_ok=True)
    return ids_today, ids_yesterday, ids_new, data_today, data_new


def dprint(*a, **k):
    if DEBUG:
        print(*a, **k)


# -------- Utilidades de ficheros (IDs) --------
def safe_read_ids_csv(ruta):
    ids = []
    if os.path.exists(ruta):
        try:
            with open(ruta, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and row[0]:
                        ids.append(row[0].strip())
        except Exception:
            return []
    return ids


def safe_write_ids_csv(ruta, ids_list):
    try:
        with open(ruta, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for _id in ids_list:
                writer.writerow([_id])
    except Exception:
        pass


def backup_to_yesterday(ids_today_path, ids_yesterday_path):
    """Copia el último today a yesterday como respaldo. No borra si no existe."""
    try:
        if os.path.exists(ids_today_path):
            shutil.copyfile(ids_today_path, ids_yesterday_path)
        else:
            # Asegura fichero vacío para coherencia
            safe_write_ids_csv(ids_yesterday_path, [])
    except Exception:
        pass


def canonicalize_fotocasa_url(url: str) -> str:
    """
    Devuelve la URL canónica sin query ni fragment.
    Normaliza el esquema/host a https://www.fotocasa.es
    """
    try:
        p = urlparse(url)
        if not p.netloc:
            # convertir ruta relativa a absoluta
            url = "https://www.fotocasa.es" + (
                url if url.startswith("/") else f"/{url}"
            )
            p = urlparse(url)
        clean = p._replace(
            scheme="https", netloc="www.fotocasa.es", params="", query="", fragment=""
        )
        return urlunparse(clean).rstrip("/")
    except Exception:
        return url.split("?")[0].split("#")[0].rstrip("/")


# -------- Utilidades de ficheros (Datos) --------
def safe_read_df_csv(ruta):
    if os.path.exists(ruta):
        try:
            return pd.read_csv(ruta)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()


def safe_write_df_csv(ruta, df):
    try:
        df_out = df.copy()
        df_out = df_out.where(pd.notnull(df_out), None)
        df_out.to_csv(ruta, index=False)
    except Exception:
        pass


# ---------- Utilidades de texto y limpieza ---------- #
def _clean_num(value: str) -> Optional[float]:
    """
    Devuelve un float a partir de una cadena numérica con separadores o símbolos.
    """
    if not value:
        return None
    value = re.sub(r"[^\d,\.]", "", value)
    value = value.replace(".", "").replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return None


def _get_li_value(li_elements: List, label: str) -> Optional[str]:
    """
    Busca el <li> cuyo <strong> empieza por `label` y devuelve su valor limpio.
    """
    for li in li_elements:
        strong = li.find("strong")
        if strong and label.lower() in strong.text.lower():
            return li.get_text(" ", strip=True).replace(strong.text, "").strip()
    return None


# ---------- Lectura de enlaces ---------- #
def read_links_from_csv(csv_path: str) -> list:
    """
    Lee los enlaces de propiedades desde un CSV y devuelve una lista de URLs.
    """
    try:
        df_links = pd.read_csv(csv_path)
        if "url" not in df_links.columns:
            raise ValueError(json.dumps({"message": "'url' column not found in CSV"}))
        links = df_links["url"].dropna().tolist()
        if not links:
            raise ValueError(json.dumps({"message": "No links found in the CSV file"}))
        return links
    except FileNotFoundError:
        raise FileNotFoundError(json.dumps({"message": f"CSV not found: {csv_path}"}))
    except Exception as e:
        raise Exception(json.dumps({"message": f"Unexpected CSV read error: {e}"}))


# ---------- Helpers de interacción ---------- #
def _click_map_tab(driver) -> None:
    """
    Intenta hacer clic en la pestaña 'Mapa' para que se carguen las coordenadas.

    - Si la pestaña no está o no es clicable, NO lanza excepción ni
      imprime logs; simplemente retorna y deja que _get_coordinates
      trate el caso sin coordenadas.
    """
    try:
        tab = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "li#liMapa a"))
        )
        driver.execute_script("arguments[0].click();", tab)
    except Exception:
        # Silencioso: cualquier problema lo ignora
        pass


# ----------  Coordenadas (tolerante a errores)  ---------- #
def _get_coordinates(driver) -> Tuple[Optional[float], Optional[float]]:
    """
    Hace clic en la pestaña “Mapa” y devuelve (lat, lon).

    - Si no se pueden obtener las coordenadas por cualquier motivo
      (pestaña no clicable, #mapa ausente, atributos vacíos, etc.),
      retorna (None, None) sin log ni excepción.
    """
    try:
        _click_map_tab(driver)

        mapa = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.ID, "mapa"))
        )
        lat = mapa.get_attribute("data-lat")
        lon = mapa.get_attribute("data-lng")
        if lat and lon:
            return float(lat), float(lon)
    except Exception:
        # Silencioso: cualquier problema implica no devolver coordenadas
        pass

    return None, None


# ---------- Extracción principal de una propiedad ---------- #
def process_property(driver, url: str) -> Optional[Dict[str, str]]:
    """
    Extrae metadatos de una página de propiedad.
    Si las coordenadas no se obtienen, ignora lat/lon en la salida.
    """
    try:
        driver.get(url)
        print(f"Scraping -> {url}")

        _handle_cookies(driver)  # banner GDPR (solo 1ª vez)

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "article#detalle"))
        )

        # ---------- Coordenadas (pueden ser None, None) ----------
        lat, lon = _get_coordinates(driver)

        # ---------- Parseo con BeautifulSoup ----------
        soup = BeautifulSoup(driver.page_source, "html.parser")

        header = soup.select_one("div.headerTitulo")
        reference = header.select_one("p span").get_text(strip=True) if header else None
        price_text = (
            header.select_one("p.precio").get_text(strip=True).split(":")[-1]
            if header
            else None
        )

        description_block = soup.select_one("#descripcionFicha p")
        description = (
            description_block.get_text(" ", strip=True) if description_block else None
        )

        details_ul = soup.select_one("div.detallesFicha ul")
        li_elements = details_ul.find_all("li") if details_ul else []

        province = _get_li_value(li_elements, "Provincia")
        city = _get_li_value(li_elements, "Población")
        zone = _get_li_value(li_elements, "Zona")
        property_type = _get_li_value(li_elements, "Tipo de propiedad")
        operation_type = _get_li_value(li_elements, "Tipo de operación")
        rooms = _get_li_value(li_elements, "Habitaciones")
        baths = _get_li_value(li_elements, "Baños")
        sup_usable = _get_li_value(li_elements, "Sup. Útil")
        sup_built = _get_li_value(li_elements, "Sup. Construida")

        # Conversión numérica
        price_eur = _clean_num(price_text)
        rooms_num = _clean_num(rooms)
        baths_num = _clean_num(baths)
        sup_usable_m2 = _clean_num(sup_usable)
        sup_built_m2 = _clean_num(sup_built)

        data = {
            "reference": reference,
            "precio_eur": price_eur,
            "provincia": province,
            "ciudad": city,
            "zona": zone,
            "tipo_de_propiedad": property_type,
            "tipo_de_operacion": operation_type,
            "habitaciones": rooms_num,
            "baños": baths_num,
            "superficie_usable_m2": sup_usable_m2,
            "superficie_construida_m2": sup_built_m2,
            "descripcion": description,
            "url": url,
            "fecha_inclusion": datetime.today().strftime("%Y-%m-%d"),
        }

        # Sólo añadimos coordenadas si existen
        if lat is not None and lon is not None:
            data["lat"] = lat
            data["lon"] = lon

        return data

    except Exception as exc:
        print(json.dumps({"message": f"Property scraping error: {exc}"}))
        return None


def parsear_inmueble(id_inmueble_url, browser, first_run):
    """
    Toma la URL canónica y usa process_property(browser, url) para extraer datos.
    Devuelve (df_row, dict_row) con el mismo esquema que antes.
    """
    import pandas as pd
    from datetime import datetime

    try:
        url = str(id_inmueble_url).strip()
        data = process_property(browser, url)  # <- fuente única de verdad

        if not data:
            return pd.DataFrame(), {}

        # Construimos campos compuestos
        localizacion = ", ".join(
            [
                p
                for p in [data.get("zona"), data.get("ciudad"), data.get("provincia")]
                if p
            ]
        )
        html = browser.page_source
        soup = bs(html, "lxml")
        details_ul = soup.select_one("div.detallesFicha ul")
        li_elements = details_ul.find_all("li") if details_ul else []

        operation_type = _get_li_value(li_elements, "Tipo de operación")
        metros_cuadrados = (
            data.get("superficie_construida_m2")
            if data.get("superficie_construida_m2") is not None
            else data.get("superficie_usable_m2")
        )

        # Mapeo al esquema original
        casas = {
            "id_inmueble": url,
            "link_inmueble": url,
            "titulo": (
                data.get("reference")
                or (
                    f"{data.get('tipo_de_propiedad') or ''} en {data.get('zona') or data.get('ciudad') or ''}".strip()
                )
            ),
            "localizacion": localizacion,
            "precio": data.get("precio_eur"),
            "metros_cuadrados": metros_cuadrados,
            "habitaciones": data.get("habitaciones"),
            "baños": data.get("baños"),
            "zona": data.get("zona"),
            "tipo_de_operacion": operation_type,
            "fecha_inclusion": data.get("fecha_inclusion")
            or datetime.today().strftime("%Y-%m-%d"),
        }

        # Si process_property añadió coordenadas, las conservamos
        if "lat" in data and data["lat"] is not None:
            casas["lat"] = data["lat"]
        if "lon" in data and data["lon"] is not None:
            casas["lon"] = data["lon"]

        df_casas = pd.DataFrame([casas])
        return df_casas, casas

    except Exception:
        # Mantenemos un fallback silencioso como antes
        import pandas as pd

        return pd.DataFrame(), {}


def build_browser():
    """
    Crea un Chrome endurecido para scraping:
      - pageLoadStrategy 'eager' para cortar carga de recursos tardíos
      - Imágenes y fuentes desactivadas
      - Memoria compartida y GPU desactivadas
      - Timeouts base configurados
    """
    opts = uc.ChromeOptions()
    opts.page_load_strategy = "eager"
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-features=NetworkService,NetworkServiceInProcess")
    opts.add_argument("--no-first-run")
    opts.add_argument("--disable-background-networking")
    opts.add_argument("--disable-background-timer-throttling")
    opts.add_argument("--disable-renderer-backgrounding")
    # Headless opcional si lo necesitas:
    # opts.add_argument("--headless=new")

    # Desactivar imágenes y fuentes para ahorrar RAM
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.managed_default_content_settings.fonts": 2,
        "profile.default_content_setting_values.images": 2,
        "profile.managed_default_content_settings.stylesheets": 1,
    }
    opts.add_experimental_option("prefs", prefs)

    browser = uc.Chrome(options=opts)
    browser.set_page_load_timeout(25)
    browser.set_script_timeout(20)
    return browser


# ----------  Gestión del banner de cookies  ---------- #
def _handle_cookies(driver) -> None:
    """
    Hace clic en “Rechazar todo” si aparece el banner GDPR.
    Se ejecuta solo la primera vez porque marca driver._cookies_rejected = True.

    - Es idempotente: si ya se ha ejecutado o no hay banner, retorna sin error.
    - Nunca lanza excepciones que paren el scraping.
    """
    # ¿Ya rechazamos cookies en esta sesión?
    if getattr(driver, "_cookies_rejected", False):
        return

    try:
        banner = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.gdprcookie"))
        )
        for btn in banner.find_elements(By.TAG_NAME, "button"):
            if "rechazar" in btn.text.lower():
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.7)  # pequeño delay para cierre del banner
                break
    except Exception:
        # Silencioso: si no hay banner o no se puede clicarlo, continuamos
        pass
    finally:
        # Marcar que ya no hace falta volver a comprobar
        setattr(driver, "_cookies_rejected", True)


# ---------- Función principal ---------- #
import os
import json
import pandas as pd
from typing import Optional, Tuple, List, Dict


# ------------------------------------------------------------------ #
# -------------------  utilidades de ficheros  ---------------------- #
# ------------------------------------------------------------------ #
def _backup_previous_file(
    base_path_no_ext: str,
) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    1. Locate the **most-recent** CSV whose name starts with ``base_path_no_ext``
       and **does not contain** the substring ``"new"`` (case-insensitive) **nor**
       end in ``"_old.csv"``.
       This is the file we will back up.
    2. Rename that file to ``<base_path_no_ext>_old.csv`` (overwriting any prior
       copy) and load it into a DataFrame for later diffs.
    3. Delete **every** other CSV and XLSX that starts with the same base path,
       including those that contain the word “new”; the only survivor is the
       freshly created “_old.csv”.

    Returns
    -------
    (df_old, backup_path)
        - df_old      : ``pandas.DataFrame`` with the contents of the backed-up
                        CSV, or ``None`` if no backup was made.
        - backup_path : Absolute path of the new «_old.csv», or ``None`` if no
                        backup was required.

    Errors
    ------
    Any unexpected exception is re-raised as ``Exception`` whose payload is a
    JSON string with a single key ``"message"`` so the frontend can display a
    concise, user-actionable error.
    """
    try:
        # --------  search patterns -------- #
        pattern_csv = f"{base_path_no_ext}*.csv"
        pattern_xlsx = f"{base_path_no_ext}*.xlsx"

        # --------  candidates for backup (exclude 'new' & '_old') -------- #
        csv_candidates = [
            p
            for p in glob.glob(pattern_csv)
            if not p.lower().endswith("_old.csv")
            and "new" not in os.path.basename(p).lower()
        ]

        if not csv_candidates:  # nothing to back up
            return None, None

        # --------  pick the latest candidate -------- #
        latest_csv = max(csv_candidates, key=os.path.getmtime)

        # Read the DataFrame before renaming
        df_old = pd.read_csv(latest_csv)

        # --------  build destination path (_old) -------- #
        backup_path = f"{base_path_no_ext}_old.csv"

        # Remove any previous _old to avoid duplicates
        if os.path.exists(backup_path):
            os.remove(backup_path)

        os.replace(latest_csv, backup_path)  # atomic move

        # --------  delete *all* other CSVs -------- #
        for f in glob.glob(pattern_csv):
            if f != backup_path and os.path.exists(f):
                os.remove(f)

        # --------  delete *all* XLSXs -------- #
        for f in glob.glob(pattern_xlsx):
            if os.path.exists(f):
                os.remove(f)

        return df_old, backup_path

    except Exception as exc:
        raise Exception(json.dumps({"message": f"Backup error: {exc}"}))


def merge_unique_ordered(list_a, list_b):
    """
    Une dos listas preservando el orden de aparición, eliminando duplicados.
    Devuelve una lista de strings.
    """
    vistos = set()
    out = []
    for v in list_a + list_b:
        s = str(v)
        if s not in vistos:
            vistos.add(s)
            out.append(s)
    return out


def _save_new_listings(
    df_new: pd.DataFrame,
    df_old: Optional[pd.DataFrame],
    base_path: str,
) -> None:
    """
    Guarda en CSV y XLSX los registros de `df_new` que no existen en `df_old`.

    - Si `df_old` es None/empty → exporta todo `df_new`.
    - Si no hay diferencias → no lanza excepción, sólo informa por consola.
    """
    try:
        # Caso sin histórico ─ exportamos todo
        if df_old is None or df_old.empty:
            _export_dataframe(df_new, base_path)
            print(f"No previous data found → exported full scrape ({len(df_new)})")
            return

        # Detectar diferencias
        key = "reference" if "reference" in df_new.columns else "url"
        df_diff = df_new[~df_new[key].isin(df_old.get(key, []))]

        # Si no hay inmuebles nuevos, salimos de forma limpia
        if df_diff.empty:
            print("No new listings detected — files not generated")
            return

        # Exportar sólo los nuevos
        _export_dataframe(df_diff, base_path)
        print(f"Detected {len(df_diff)} new listings")
    except Exception as exc:
        raise Exception(json.dumps({"message": f"Saving new listings error: {exc}"}))


import os
import json
import pandas as pd
from typing import Optional, Tuple, List, Dict


LISTING_URL_TEMPLATE = "https://www.picoblanes.com/results/?id_tipo_operacion={op}&type=&dt=&dormitorios_min=&precio_max="
OPERATION_TYPES: tuple[int, ...] = (
    1,
    2,
    4,
)  # 1 = venta, 2 = alquiler, 4 = Alquiler opciónn a compra

SLEEP_BETWEEN_REQUESTS = 1.5  # s: juega limpio
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    )
}


# ――― Utilidades HTTP / parseo ―――─────────────────────────────────────────────
def fetch_html(url: str) -> BeautifulSoup:
    """Devuelve el árbol DOM (BeautifulSoup) de la página indicada."""
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def extract_property_urls(soup: BeautifulSoup, base_url: str) -> List[str]:
    """
    Extrae las URLs absolutas de los inmuebles de la página.
    El enlace está en el atributo data-url del <div class="venta">.
    """
    result: list[str] = []
    for div in soup.select("div[data-url]"):
        relative = div["data-url"].strip()
        result.append(urljoin(base_url, relative))
    return result
    return result


def find_next_page(soup: BeautifulSoup, current_url: str) -> Optional[str]:
    """
    Devuelve la URL de la siguiente página, o None si no existe.
    Se basa en <a class="next"> que no esté 'disabled'.
    """
    next_link = soup.select_one("a.next:not(.disabled)")
    if not next_link:
        return None

    href = next_link.get("href")
    if not href:
        return None
    return urljoin(current_url, href)


def _add_date_suffix(path_no_ext: str) -> str:
    """
    Devuelve el mismo `path_no_ext` con sufijo _YYYYMMDD si el nombre
    base contiene la subcadena “inmueble” (indistinto de mayúsculas).

    Parameters
    ----------
    path_no_ext : str
        Ruta completa sin extensión, p.e. '.../pico_blanes_inmuebles'

    Returns
    -------
    str
        Mismo path, potencialmente con _YYYYMMDD añadido.
    """
    base_name = os.path.basename(path_no_ext).lower()
    if "inmueble" in base_name:
        today = datetime.now().strftime("%Y%m%d")
        return f"{path_no_ext}_{today}"
    return path_no_ext


def scrape_pico_blanes_ids(start_url, seen_props):
    seen_pages: set[str] = set()
    new_props: list[str] = []

    url = start_url
    while url and url not in seen_pages:
        print(f"[+] Descargando {url}", file=sys.stderr)
        seen_pages.add(url)

        soup = fetch_html(url)
        props = extract_property_urls(soup, url)

        for u in props:
            if u not in seen_props:
                seen_props.add(u)
                new_props.append(u)

        print(
            f"    · {len(new_props)} nuevas — total acumulado: {len(seen_props)}",
            file=sys.stderr,
        )

        url = find_next_page(soup, url)
        time.sleep(SLEEP_BETWEEN_REQUESTS)


# ------------------------------------------------------------------ #
# --------------- utilidades de exportación múltiples ---------------#
# ------------------------------------------------------------------ #
def _export_dataframe(df: pd.DataFrame, base_path: str) -> None:
    """
    Exporta `df` a dos formatos: CSV y XLSX con filtros activados.

    • Si `base_path` tiene “inmueble” en el nombre ⇒ se añade la fecha
      (YYYYMMDD) antes de la extensión.
    • Los backups *_old.* NO se modifican nunca.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame a exportar.
    base_path : str
        Ruta base sin extensión (p.e. 'data/pico_blanes_inmuebles').

    Errores controlados
    -------------------
    - Si el DataFrame está vacío se lanza ValueError.
    - Cualquier IOError se envuelve y propaga con un mensaje JSON.
    """
    if df.empty:
        raise ValueError(
            json.dumps({"message": "DataFrame is empty — nothing to export"})
        )

    # ----------- rutas de salida -----------
    dated_base = _add_date_suffix(base_path)
    csv_path = f"{dated_base}.csv"
    xlsx_path = f"{dated_base}.xlsx"

    try:
        # ----------- CSV -----------
        df.to_csv(csv_path, index=False)

        # ----------- Excel -----------
        df.to_excel(xlsx_path, index=False)

        # convertimos el rango en tabla con autofiltro
        wb = load_workbook(xlsx_path)
        ws = wb.active
        last_col_letter = get_column_letter(ws.max_column)
        last_row = ws.max_row
        table_ref = f"A1:{last_col_letter}{last_row}"

        table = Table(displayName="DataTable", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        ws.freeze_panes = "A2"

        wb.save(xlsx_path)
        wb.close()

        print(f"Saved    -> {csv_path} & {xlsx_path} (filters enabled)")

    except Exception as exc:
        raise IOError(json.dumps({"message": f"File export error: {exc}"}))


# ------------------------------------------------------------------ #
# -------------------  utilidades de ficheros  ---------------------- #
# ------------------------------------------------------------------ #
def _save_new_listings(
    df_new: pd.DataFrame,
    df_old: Optional[pd.DataFrame],
    base_path: str,
) -> None:
    """
    Guarda los registros nuevos respecto a `df_old` en CSV y XLSX.

    Parameters
    ----------
    df_new : pd.DataFrame
        DataFrame con todos los inmuebles scrapeados.
    df_old : pd.DataFrame or None
        Histórico previo. Si None o vacío, se consideran todos como nuevos.
    base_path : str
        Ruta base sin extensión (p.e. 'data/properties_new').
    """
    try:
        if df_old is None or df_old.empty:
            _export_dataframe(df_new, base_path)
            print(f"No previous data found → exported full scrape ({len(df_new)})")
            return

        key = "reference" if "reference" in df_new.columns else "url"
        df_diff = df_new[~df_new[key].isin(df_old.get(key, []))]
        _export_dataframe(df_diff, base_path)
        print(f"Detected {len(df_diff)} new listings")
    except Exception as exc:
        raise Exception(json.dumps({"message": f"Saving new listings error: {exc}"}))


# ------------------------------------------------------------------ #
# --------------------------  main  -------------------------------- #
# ------------------------------------------------------------------ #
def main():
    # Rutas
    ids_today_file, ids_yesterday_file, ids_new_file, data_today_file, data_new_file = (
        paths_for()
    )
    dprint(f"[INIT] BASE_DIR={BASE_DIR}")
    dprint(f"[FILES] ids_today={ids_today_file}")
    dprint(f"[FILES] ids_yesterday={ids_yesterday_file}")
    dprint(f"[FILES] data_today={data_today_file}")

    # 0) Ayer
    backup_to_yesterday(ids_today_file, ids_yesterday_file)

    # 1) Cargar data_today previa
    df_today_prev = safe_read_df_csv(data_today_file)
    if df_today_prev.empty:
        dprint("[DATA_TODAY] VACÍO o no legible.")
        existing_ids_in_data = set()
    else:
        if "id_inmueble" not in df_today_prev.columns:
            dprint(
                "[DATA_TODAY] ¡Falta columna 'id_inmueble'! Columns=",
                list(df_today_prev.columns),
            )
            existing_ids_in_data = set()
        else:
            existing_ids_in_data = set(
                df_today_prev["id_inmueble"].astype(str).tolist()
            )
            dprint(
                f"[DATA_TODAY] filas={len(df_today_prev)} ids_unicos={len(existing_ids_in_data)} sample={list(existing_ids_in_data)[:5]}"
            )

    # 2) Navegador y scrape

    all_property_urls: set[str] = set()
    for op in OPERATION_TYPES:
        start_url = LISTING_URL_TEMPLATE.format(op=op)
        scrape_pico_blanes_ids(start_url, all_property_urls)

    ids_hoy = list(all_property_urls)

    dprint(f"[HOY] ids_hoy={len(ids_hoy)} sample={ids_hoy[:10]}")

    # 4) Persistir ids_today
    safe_write_ids_csv(ids_today_file, ids_hoy)

    # 5) Calcular "nuevos": en ids_hoy y NO en el CSV de inmuebles (data_today)
    ids_nuevos = [str(i) for i in ids_hoy if str(i) not in existing_ids_in_data]
    safe_write_ids_csv(ids_new_file, ids_nuevos)

    # Nuevo navegador para parseo
    browser = build_browser()

    # Parsear solo nuevos con reintentos y reciclado del driver
    first_run = True
    df_new_list = []
    print(f"Parseando {len(ids_nuevos)} nuevos inmuebles Fotocasa...")

    RECYCLE_EVERY = 100  # recicla el driver cada N fichas
    CLEAR_STATE_EVERY = 5  # limpia cache/cookies cada N fichas
    MAX_RETRIES_PER_URL = 2  # reintentos por ficha

    for idx, url in enumerate(ids_nuevos, start=1):
        # Limpieza periódica de estado para frenar el crecimiento de memoria
        if idx % CLEAR_STATE_EVERY == 0:
            clear_browser_state(browser, clear_cache=True, clear_cookies=True)

        # Reciclado periódico del driver para cortar fugas
        if idx % RECYCLE_EVERY == 0:
            try:
                browser.quit()
            except Exception:
                pass
            browser = build_browser()
            first_run = True  # tras reciclar, volvemos a aceptar cookies si salen

        success = False
        for attempt in range(1, MAX_RETRIES_PER_URL + 1):
            try:
                df_i, _ = parsear_inmueble(url, browser, first_run)
                if not df_i.empty:
                    df_new_list.append(df_i)
                    success = True
                    break
            except WebDriverException:
                # Driver en estado malo: reciclar y reintentar
                try:
                    browser.quit()
                except Exception:
                    pass
                browser = build_browser()
                first_run = True
            except Exception:
                # Fallo blando; reintento
                pass
            # backoff simple

        first_run = False

        if not success:
            # Registro mínimo en consola, no se escribe fila vacía
            print(f"[WARN] No se pudo parsear: {url}")

    # Cierre del navegador
    try:
        browser.quit()
    except Exception:
        pass

    # Consolidación y escritura segura
    df_new = (
        pd.concat(df_new_list, ignore_index=True)
        if df_new_list
        else pd.DataFrame(
            columns=[
                "id_inmueble",
                "titulo",
                "localizacion",
                "municipio",
                "precio",
                "precio_bajada",
                "metros_cuadrados",
                "habitaciones",
                "baños",
                "consumo_energia",
                "emisiones_energia",
                "descripcion",
                "breadcrumb",
                "fotos_total",
                "ref_catastral",
                "edificabilidad_m2techo",
                "parcela_m2",
                "parcela_min_m2",
                "fachada_min_m",
                "altura_max_m",
                "sector_urbanistico",
                "link_inmueble",
                "link_inmueble_siguiente",
                "anunciante",
                "zona",
                "tipo_de_operacion",
            ]
        )
    )

    safe_write_df_csv(data_new_file, df_new)

    # today consistente
    if not df_today_prev.empty:
        df_today_kept = df_today_prev[
            df_today_prev["id_inmueble"].astype(str).isin(set(map(str, ids_hoy)))
        ].copy()
    else:
        df_today_kept = pd.DataFrame(columns=df_new.columns)

    df_today = pd.concat([df_today_kept, df_new], ignore_index=True)
    if not df_today.empty:
        df_today = df_today.drop_duplicates(subset=["id_inmueble"], keep="last")

    safe_write_df_csv(data_today_file, df_today)
    print(
        f"Escritos: {data_new_file} ({len(df_new)} filas) y {data_today_file} ({len(df_today)} filas)."
    )


if __name__ == "__main__":
    main()
